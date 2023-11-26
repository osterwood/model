
class Location:

    def __init__(self, name, homes, heating_fuels):
        self.name = name
        self.homes = homes
        self.heating_fuels = heating_fuels

    def likely_heating_fuel(self):
        key = max(self.heating_fuels, key=self.heating_fuels.get)
        value = self.heating_fuels[key]
        return (key, value)

    @classmethod
    def parse(cls, row):

        name = row[0].value.strip()
        homes = float(row[1].value)

        fuels = dict()

        for idx, fuel in [
            [2, 'electricity'],
            [4, 'natural_gas'],
            [6, 'propane'],
            [8, 'fuel_oil']
        ]:
            value = row[idx].value

            if value == 'N' or value == 'Q':
                fuels[fuel] = 0.0
            else:
                fuels[fuel] = value / homes

        fuels['other'] = 1.0 - sum(fuels.values())

        return Location(name, homes, fuels)



if __name__ == '__main__':
    from openpyxl import load_workbook
    import os, inspect
    
    lib_folder = os.path.join(os.path.split(inspect.getfile( inspect.currentframe() ))[0], '..', 'lib')
    wb = load_workbook(filename = '{}/../data/State Space Heating Fuels.xlsx'.format(lib_folder))
    in_header = True
    in_footer = False

    states = []

    for row in wb['data'].iter_rows():
        if in_header and row[0].value == 'All homes':
            in_header = False
            continue

        if len(states) >= 50:
            break

        if not in_header and row[0].value is not None:
            state = Location.parse(row)
            states.append(state)

    for state in states:
        print(state.name, state.likely_heating_fuel())
